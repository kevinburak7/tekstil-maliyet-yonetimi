"""Sprint A+B: formül, validasyon ve DB birim testleri."""
import os
import tempfile
import unittest

import maliyet
from maliyet import (
    FIRE_ORANI_VARSAYILAN,
    ValidationError,
    Veritabani,
    maliyet_hesapla,
    parse_float,
    recete_toplam,
)


KURLAR = {"TL": 1.0, "USD": 34.0, "EUR": 37.0}


class TestParseFloat(unittest.TestCase):
    def test_virgul(self):
        self.assertEqual(parse_float("12,5"), 12.5)

    def test_bos(self):
        with self.assertRaises(ValidationError):
            parse_float("  ")

    def test_gecersiz(self):
        with self.assertRaises(ValidationError):
            parse_float("abc")


class TestMaliyetHesapla(unittest.TestCase):
    def test_boya(self):
        item = {"miktar": 2, "fiyat": 100, "para": "TL", "birim": "% (Boya)"}
        self.assertAlmostEqual(maliyet_hesapla(item, "Boya", 1, KURLAR), 2.4)

    def test_apre(self):
        item = {"miktar": 10, "fiyat": 50, "para": "TL", "birim": "g/l (Apre)"}
        self.assertAlmostEqual(maliyet_hesapla(item, "Apre", 80, KURLAR), 48.0)

    def test_kimyasal_gl(self):
        item = {"miktar": 5, "fiyat": 20, "para": "TL", "birim": "g/l"}
        self.assertAlmostEqual(maliyet_hesapla(item, "Kimyasal", 8, KURLAR), 0.96)

    def test_kimyasal_yuzde(self):
        item = {"miktar": 2, "fiyat": 100, "para": "TL", "birim": "%"}
        self.assertAlmostEqual(maliyet_hesapla(item, "Kimyasal", 8, KURLAR), 2.4)

    def test_usd_kur(self):
        item = {"miktar": 1, "fiyat": 1, "para": "USD", "birim": "% (Boya)"}
        self.assertAlmostEqual(maliyet_hesapla(item, "Boya", 1, KURLAR), 0.408)

    def test_flotte_sifir(self):
        item = {"miktar": 1, "fiyat": 1, "para": "TL", "birim": "g/l"}
        with self.assertRaises(ValidationError):
            maliyet_hesapla(item, "Kimyasal", 0, KURLAR)

    def test_recete_toplam(self):
        items = [
            {"miktar": 2, "fiyat": 100, "para": "TL", "birim": "%"},
            {"miktar": 1, "fiyat": 50, "para": "TL", "birim": "%"},
        ]
        self.assertAlmostEqual(recete_toplam(items, "Kimyasal", 10, KURLAR), 3.0)

    def test_fire_orani_sabiti(self):
        self.assertEqual(FIRE_ORANI_VARSAYILAN, 1.2)

    def test_ozel_fire_orani(self):
        item = {"miktar": 2, "fiyat": 100, "para": "TL", "birim": "% (Boya)"}
        # 2 * 100 / 100 * 1.5 = 3.0
        self.assertAlmostEqual(maliyet_hesapla(item, "Boya", 1, KURLAR, 1.5), 3.0)

    def test_negatif_miktar(self):
        item = {"miktar": -1, "fiyat": 100, "para": "TL", "birim": "% (Boya)"}
        with self.assertRaises(ValidationError):
            maliyet_hesapla(item, "Boya", 1, KURLAR)

    def test_bilinmeyen_para(self):
        item = {"miktar": 1, "fiyat": 100, "para": "GBP", "birim": "% (Boya)"}
        with self.assertRaises(ValidationError):
            maliyet_hesapla(item, "Boya", 1, KURLAR)

    def test_parse_pozitif(self):
        from maliyet import parse_pozitif

        self.assertEqual(parse_pozitif("3"), 3.0)
        with self.assertRaises(ValidationError):
            parse_pozitif("0")
        with self.assertRaises(ValidationError):
            parse_pozitif("-2")



class TestVeritabani(unittest.TestCase):
    def setUp(self):
        import tekstil_maliyet.constants as constants

        self._fd, self.path = tempfile.mkstemp(suffix=".db")
        os.close(self._fd)
        self._old = constants.DB_YOLU
        constants.DB_YOLU = self.path
        self.db = Veritabani()

    def tearDown(self):
        import tekstil_maliyet.constants as constants

        self.db.conn.close()
        constants.DB_YOLU = self._old
        try:
            os.remove(self.path)
        except OSError:
            pass

    def test_kaydet_getir_guncelle_sil(self):
        icerik = [
            {"ad": "Asit", "miktar": 2.0, "birim": "g/l", "fiyat": 10.0, "para": "TL"}
        ]
        rid = self.db.kaydet("Kimyasal", "Test Recete", 8.0, icerik, fire_orani=1.35)
        self.assertIsInstance(rid, int)

        rec = self.db.getir_by_id(rid)
        self.assertEqual(rec["isim"], "Test Recete")
        self.assertAlmostEqual(rec["fire_orani"], 1.35)
        self.assertEqual(len(rec["icerik"]), 1)
        self.assertEqual(rec["icerik"][0]["ad"], "Asit")

        yeni = [
            {"ad": "Tuz", "miktar": 1.0, "birim": "%", "fiyat": 5.0, "para": "USD"}
        ]
        self.assertTrue(
            self.db.guncelle(rid, "Kimyasal", "Test Recete 2", 10.0, yeni, fire_orani=1.1)
        )
        rec2 = self.db.getir_by_id(rid)
        self.assertEqual(rec2["isim"], "Test Recete 2")
        self.assertEqual(rec2["parametre"], 10.0)
        self.assertAlmostEqual(rec2["fire_orani"], 1.1)
        self.assertEqual(rec2["icerik"][0]["ad"], "Tuz")

        self.assertEqual(self.db.son_id_by_isim("Kimyasal", "Test Recete 2"), rid)
        self.assertTrue(self.db.sil_by_id(rid))
        self.assertIsNone(self.db.getir_by_id(rid))

    def test_katalog_crud(self):
        kid = self.db.katalog_ekle("Asit A", "g/l", 12.5, "USD")
        self.assertIsInstance(kid, int)
        liste = self.db.katalog_listele()
        self.assertEqual(len(liste), 1)
        self.assertEqual(liste[0]["ad"], "Asit A")
        self.assertTrue(
            self.db.katalog_guncelle(kid, "Asit B", "%", 8.0, "TL", aktif=True)
        )
        urun = self.db.katalog_getir_by_ad("Asit B")
        self.assertEqual(urun["fiyat"], 8.0)
        self.assertTrue(self.db.katalog_sil(kid))
        self.assertEqual(self.db.katalog_listele(sadece_aktif=False), [])

    def test_katalog_pasif(self):
        kid = self.db.katalog_ekle("Pasif Urun", "g/l", 1.0, "TL")
        self.assertTrue(
            self.db.katalog_guncelle(kid, "Pasif Urun", "g/l", 1.0, "TL", aktif=False)
        )
        self.assertEqual(self.db.katalog_listele(sadece_aktif=True), [])
        hepsi = self.db.katalog_listele(sadece_aktif=False)
        self.assertEqual(len(hepsi), 1)
        self.assertFalse(hepsi[0]["aktif"])


class TestExcelAktar(unittest.TestCase):
    def test_excel_dosya_olusturur(self):
        from maliyet import excel_aktar

        receteler = [
            {
                "id": 1,
                "tur": "Boya",
                "isim": "Demo",
                "parametre": 1.0,
                "fire_orani": 1.5,
                "tarih": "2026-01-01",
                "icerik": [
                    {
                        "ad": "Boy",
                        "miktar": 2.0,
                        "birim": "% (Boya)",
                        "fiyat": 100.0,
                        "para": "TL",
                    }
                ],
            }
        ]
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            excel_aktar(path, receteler, KURLAR)
            self.assertTrue(os.path.getsize(path) > 0)
            from openpyxl import load_workbook

            wb = load_workbook(path)
            self.assertIn("Ozet", wb.sheetnames)
            self.assertIn("Detay", wb.sheetnames)
            self.assertEqual(wb["Ozet"][2][2].value, "Demo")
            self.assertEqual(wb["Ozet"][2][4].value, 1.5)
            # 2*100/100*1.5 = 3.0
            self.assertAlmostEqual(float(wb["Ozet"][2][6].value), 3.0)
        finally:
            try:
                os.remove(path)
            except OSError:
                pass


class TestPdfAktar(unittest.TestCase):
    def test_pdf_dosya_olusturur(self):
        from maliyet import pdf_aktar

        receteler = [
            {
                "id": 1,
                "tur": "Boya",
                "isim": "Demo Boya",
                "parametre": 1.0,
                "fire_orani": 1.2,
                "tarih": "2026-01-01",
                "icerik": [
                    {
                        "ad": "Boy",
                        "miktar": 2.0,
                        "birim": "% (Boya)",
                        "fiyat": 100.0,
                        "para": "TL",
                    }
                ],
            }
        ]
        fd, path = tempfile.mkstemp(suffix=".pdf")
        os.close(fd)
        try:
            pdf_aktar(path, receteler, KURLAR)
            self.assertTrue(os.path.getsize(path) > 100)
            with open(path, "rb") as f:
                self.assertEqual(f.read(4), b"%PDF")
        finally:
            try:
                os.remove(path)
            except OSError:
                pass


class TestKarsilastir(unittest.TestCase):
    def test_en_ucuz_ve_fark(self):
        from maliyet import karsilastir

        a = {
            "id": 1,
            "tur": "Boya",
            "isim": "Ucuz",
            "parametre": 1.0,
            "fire_orani": 1.2,
            "icerik": [
                {
                    "ad": "Boy",
                    "miktar": 1.0,
                    "birim": "% (Boya)",
                    "fiyat": 100.0,
                    "para": "TL",
                }
            ],
        }
        b = {
            "id": 2,
            "tur": "Boya",
            "isim": "Pahali",
            "parametre": 1.0,
            "fire_orani": 1.2,
            "icerik": [
                {
                    "ad": "Boy",
                    "miktar": 2.0,
                    "birim": "% (Boya)",
                    "fiyat": 100.0,
                    "para": "TL",
                }
            ],
        }
        # a: 1.2, b: 2.4
        sonuc = karsilastir([a, b], KURLAR)
        self.assertEqual(sonuc["en_ucuz_id"], 1)
        self.assertAlmostEqual(sonuc["en_ucuz_maliyet"], 1.2)
        pahali = next(o for o in sonuc["ozetler"] if o["id"] == 2)
        self.assertAlmostEqual(pahali["fark_tl"], 1.2)
        self.assertAlmostEqual(pahali["fark_yuzde"], 100.0)
        self.assertEqual(len(sonuc["urun_satirlari"]), 1)
        self.assertEqual(sonuc["urun_satirlari"][0]["ad"], "Boy")

    def test_en_az_iki(self):
        from maliyet import karsilastir

        with self.assertRaises(ValueError):
            karsilastir(
                [
                    {
                        "id": 1,
                        "tur": "Boya",
                        "isim": "X",
                        "parametre": 1,
                        "icerik": [],
                    }
                ],
                KURLAR,
            )


if __name__ == "__main__":
    unittest.main()
